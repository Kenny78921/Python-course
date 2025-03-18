from openpyxl import load_workbook

wb = load_workbook(filename="main.xlsx")

ws = wb.active
for x in range(1, 10):
    ws["A" + str(x)] = "Hola"

wb.save(filename="main.xlsx")
